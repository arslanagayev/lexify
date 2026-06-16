"""
Parse uploaded CSV / Excel files into word rows for bulk import.
Expected columns: `word` (required), `notes` (optional). Case-insensitive headers.
"""
from __future__ import annotations

import csv
import io
from typing import Optional

MAX_WORDS = 50


def _norm_header(h: Optional[str]) -> str:
    return (h or "").strip().lower()


def parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    headers = [_norm_header(h) for h in rows[0]]
    # Support header-less single-column files too
    if "word" not in headers:
        # treat every first cell as a word
        return [{"word": r[0].strip(), "notes": ""} for r in rows if r and r[0].strip()]
    w_idx = headers.index("word")
    n_idx = headers.index("notes") if "notes" in headers else None
    out = []
    for r in rows[1:]:
        if w_idx >= len(r):
            continue
        word = (r[w_idx] or "").strip()
        if not word:
            continue
        notes = (r[n_idx].strip() if n_idx is not None and n_idx < len(r) else "")
        out.append({"word": word, "notes": notes})
    return out


def parse_excel(content: bytes) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [_norm_header(str(h) if h is not None else "") for h in rows[0]]
    if "word" not in headers:
        return [
            {"word": str(r[0]).strip(), "notes": ""}
            for r in rows if r and r[0] is not None and str(r[0]).strip()
        ]
    w_idx = headers.index("word")
    n_idx = headers.index("notes") if "notes" in headers else None
    out = []
    for r in rows[1:]:
        if w_idx >= len(r) or r[w_idx] is None:
            continue
        word = str(r[w_idx]).strip()
        if not word:
            continue
        notes = ""
        if n_idx is not None and n_idx < len(r) and r[n_idx] is not None:
            notes = str(r[n_idx]).strip()
        out.append({"word": word, "notes": notes})
    return out


def parse_import_file(filename: str, content: bytes) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        rows = parse_excel(content)
    else:
        rows = parse_csv(content)
    # De-dupe within the file (case-insensitive), preserve order, cap at MAX_WORDS
    seen, deduped = set(), []
    for row in rows:
        key = row["word"].lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped[:MAX_WORDS]
